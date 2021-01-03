import openpyxl, sys

file = sys.argv[1]
wb = openpyxl.load_workbook(file)
sheet = wb.active
newSheet = wb.create_sheet(index = 0, title = 'inverted')

for x in sheet.rows:
    for y in x:
        newSheet.cell(row = y.column, column = y.row).value = y.value
wb.save('inverted_{}'.format(file))