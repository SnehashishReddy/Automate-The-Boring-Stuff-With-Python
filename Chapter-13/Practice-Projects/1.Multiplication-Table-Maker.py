import sys, openpyxl
from openpyxl.styles import Font

# Setting the table size
x = int(sys.argv[1])
# Opening the Workbook
wb = openpyxl.Workbook()
sheet = wb['Sheet']
# Setting the Bold Style
bold12Font = Font(size = 14, bold = True)

# Nested Loop to write out the spreadsheet
for a in range(0,x):
    for b in range(0,x):
        # Setting all the numbers in the 1st column
        if a == 0:
            sheet.cell(row = b+1, column = 1).font = bold12Font
            sheet.cell(row = b+1, column = 1).value = b+1
        # Setting all the numbers in the 1st row
        elif b == 0:
            sheet.cell(row = 1, column = a+1).font = bold12Font
            sheet.cell(row = 1, column = a+1).value = a+1
        # Multiplying those numbers afterwards
        else:
            activeCell = sheet.cell(row = b+1, column = a+1)
            activeCell.value = (a+1)*(b+1)
wb.save("Multiplication-Table-for-{}.xlsx".format(x))